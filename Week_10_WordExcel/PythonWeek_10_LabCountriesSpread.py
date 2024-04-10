"""Your boss has asked you to make a spreadsheet of every country in the world, and their capital cities.

There are over 200 countries in the world. So, you are going to use a Python program to create the Excel spreadsheet.

This API can provide a list of counties and their capital cities via this url:
https://country-list-1150.azurewebsites.net/api/country

(A JSON Formatter is helpful to see the structure of the response. Try the Firefox browser, or install a Chrome
browser extension like JSON Formatter or use pprint in your program.)

In your program,

Use requests to fetch the list from the API https://country-list-1150.azurewebsites.net/api/country and convert the
JSON into a Python object.
For each list item in the response: Extract each country name and capital city from the
response Uses openpxl to write the names and capital cities to a spreadsheet.
The country names should be in column
A, and their capital city in column B."""
import requests  # must do both
from openpyxl import Workbook  # import library to work with Excel documents in Python -

# this imports just the module we need

#  make request to API to get list of country info
countries_url = 'https://country-list-1150.azurewebsites.net/api/country'  # this is a list of dictionaries, with nested
# dictionaries. usually loop over list, access data by keys
country_list_response = requests.get(countries_url).json()
# descriptive variable names are important
#  create workbook - before loop~

# to have data written to it as loop runs, and we access the country and capital
# looking things up is expected and normal as programmer

country_workbook = Workbook()
country_worksheet = country_workbook.active
country_worksheet.title = 'Countries and Capital Cities'
row = 0
column = 0
#  get the country and capital city from response
for individual_country_dictionary in country_list_response:
    # print(individual_country_dictionary) # huge list, we need to break it down aka access data
    country_name = individual_country_dictionary['name']  # this accesses list of nested dictionaries to give us data
    # based on key 'name'
    # print(country_name) this will print individual country name - good for debugging and checking data
    capital_city = individual_country_dictionary['capitalCity']
    print(f'{country_name:<35} {capital_city}')  # this prints data called from list of dictionaries, format string to
    # use :<35 to add spacing in table
    row = row + 1
    column = column + 1
    country_worksheet.cell(row, 1, country_name)
    country_worksheet.cell(column, 2, capital_city)

country_workbook.save('counties_and_capitals.xlsx')
#  write county and capital city to Workbook
